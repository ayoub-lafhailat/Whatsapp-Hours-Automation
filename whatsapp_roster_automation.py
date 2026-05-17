import datetime
import time
from pathlib import Path

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By


def send_roster_to_whatsapp():
    driver = None

    try:
        date = datetime.datetime.now()
        column_index = int(date.strftime("%d")) - 1

        chrome_user_data_dir = r"C:\Users\Beheerder\AppData\Local\Google\Chrome\User Data"
        chromedriver_path = r"C:\Users\Beheerder\PycharmProjects\ROOSTER\chromedriver.exe"

        options = webdriver.ChromeOptions()
        options.add_argument(fr"user-data-dir={chrome_user_data_dir}")

        service = Service(chromedriver_path)
        driver = webdriver.Chrome(service=service, options=options)

        driver.get("https://web.whatsapp.com/")
        wait = WebDriverWait(driver, 600)

        with open("MAAND.txt", "r", encoding="utf-8") as file:
            month_name = file.read().strip()

        excel_file = f"{month_name}.xlsx"
        print(f"Using Excel file: {excel_file}")

        if not Path(excel_file).exists():
            raise FileNotFoundError(f"Excel file not found: {excel_file}")

        workbook = load_workbook(filename=excel_file)

        location_groups = {
            "GELDERMALSEN": '"Bodegraven"',
            "BEST": '"Best"',
            "ROSMALEN": '"Rosmalen"',
        }

        for sheet_name, whatsapp_group in location_groups.items():
            if sheet_name not in workbook.sheetnames:
                print(f"Sheet not found: {sheet_name}")
                continue

            sheet = workbook[sheet_name]
            print(sheet_name, whatsapp_group)

            roster_lines = []

            plan_row = sheet.max_row
            for row_index in range(1, plan_row + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                roster_lines.append(str(cell.value))

            x_arg = f"//span[contains(@title,{whatsapp_group})]"
            group_title = wait.until(
                EC.presence_of_element_located((By.XPATH, x_arg))
            )
            group_title.click()

            time.sleep(2)

            input_xpath = '//div[@class="_13NKt copyable-text selectable-text"][@data-tab="10"]'
            input_box = wait.until(
                EC.presence_of_element_located((By.XPATH, input_xpath))
            )

            message = " \n".join(roster_lines)
            print(message)

            input_box.send_keys("ROOSTER " + message + Keys.ENTER)
            time.sleep(2)

    finally:
        if driver is not None:
            driver.close()


if __name__ == "__main__":
    send_roster_to_whatsapp()